# spreadsheet_handler.py - Store spreadsheet in same directory as app with proper formatting
import pandas as pd
import os
from datetime import datetime, time as dt_time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

class SpreadsheetHandler:
    def __init__(self, config):
        self.config = config
        
        # Always use the application directory for the spreadsheet
        app_dir = os.path.dirname(os.path.abspath(__file__))
        self.file_path = os.path.join(app_dir, 'maintenance_jobs.xlsx')
        self.sheet_name = config['spreadsheet']['sheet_name']
        
        # Override config with actual path
        self.config['spreadsheet']['file_path'] = self.file_path
        
        print(f"📊 Spreadsheet path: {self.file_path}")
        
        # Initialize spreadsheet if it doesn't exist
        if not os.path.exists(self.file_path):
            print("Creating new spreadsheet in app directory...")
            self.initialize_spreadsheet()
        else:
            print("Spreadsheet already exists in app directory")
            # Ensure it has the latest columns
            self.ensure_columns_exist()
    
    def ensure_columns_exist(self):
        """Ensure all required columns exist in the spreadsheet"""
        try:
            if os.path.exists(self.file_path):
                df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
                
                # Define all required columns IN THE CORRECT ORDER
                required_columns = [
                    'Job Number',
                    'Creation Date',
                    'Start Time',
                    'End Time',
                    'Room Number',
                    'Issue',
                    'Resolution'
                ]
                
                # Check if columns are missing or in wrong order
                current_columns = list(df.columns)
                
                if current_columns != required_columns:
                    print(f"🔄 Reorganizing columns to correct order...")
                    
                    # Reorder existing columns and add missing ones
                    new_df = pd.DataFrame(columns=required_columns)
                    
                    # Copy existing data if columns exist
                    for col in required_columns:
                        if col in df.columns:
                            new_df[col] = df[col]
                        else:
                            new_df[col] = ''
                    
                    # Save with proper formatting
                    self.save_with_formatting(new_df)
                    
                    print("✅ Columns reorganized to correct order")
                    
        except Exception as e:
            print(f"⚠️  Error ensuring columns exist: {e}")
            # Reinitialize if there's an error
            self.initialize_spreadsheet()
    
    def initialize_spreadsheet(self):
        """Create a new spreadsheet with headers in app directory with proper formatting"""
        try:
            # Create DataFrame with headers IN THE CORRECT ORDER
            df = pd.DataFrame(columns=[
                'Job Number',
                'Creation Date',
                'Start Time',
                'End Time',
                'Room Number',
                'Issue',
                'Resolution'
            ])
            
            print(f"💾 Creating new spreadsheet: {self.file_path}")
            
            # Save with formatting
            self.save_with_formatting(df)
            
            print(f"✅ Spreadsheet created: {self.file_path}")
            
        except Exception as e:
            print(f"❌ Error creating spreadsheet: {e}")
            raise
    
    def save_with_formatting(self, df):
        """Save DataFrame to Excel with proper cell formatting"""
        try:
            # Make a copy to avoid modifying the original
            df_to_save = df.copy()
            
            # Ensure proper data types
            df_to_save['Job Number'] = pd.to_numeric(df_to_save['Job Number'], errors='coerce')
            df_to_save['Creation Date'] = pd.to_datetime(df_to_save['Creation Date'], errors='coerce')
            
            # Handle time columns - convert HH:MM strings to Excel time format
            time_columns = ['Start Time', 'End Time']
            for col in time_columns:
                if col in df_to_save.columns:
                    # Convert HH:MM strings to datetime.time objects
                    df_to_save[col] = df_to_save[col].apply(self.parse_time_string)
            
            # Create a Pandas Excel writer using openpyxl
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                df_to_save.to_excel(writer, sheet_name=self.sheet_name, index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[self.sheet_name]
                
                # Define styles
                header_font = Font(bold=True, size=12)
                thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'),
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                
                # Format headers
                for col_num, value in enumerate(df_to_save.columns.values, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Auto-size columns based on content
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    # Find the maximum length in the column
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set column width (with some padding)
                    adjusted_width = min(max_length + 2, 50)  # Max width of 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Set specific column widths for better readability
                worksheet.column_dimensions['A'].width = 12  # Job Number
                worksheet.column_dimensions['B'].width = 15  # Creation Date
                worksheet.column_dimensions['C'].width = 12  # Start Time
                worksheet.column_dimensions['D'].width = 12  # End Time
                worksheet.column_dimensions['E'].width = 12  # Room Number
                worksheet.column_dimensions['F'].width = 40  # Issue
                worksheet.column_dimensions['G'].width = 40  # Resolution
                
                # Apply text wrapping and borders to all data cells
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                              min_col=1, max_col=len(df_to_save.columns)):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = thin_border
                
                # Format time columns
                for col_letter in ['C', 'D']:  # Start Time and End Time columns
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet[f'{col_letter}{row}']
                        if cell.value:
                            cell.number_format = 'HH:MM'
                
                # Format date column
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'B{row}']  # Creation Date column
                    if cell.value:
                        cell.number_format = 'YYYY-MM-DD'
                
                # Freeze the header row
                worksheet.freeze_panes = 'A2'
                
                print("✅ Spreadsheet saved with proper formatting")
                
        except Exception as e:
            print(f"❌ Error saving spreadsheet with formatting: {e}")
            # Fall back to simple save
            df.to_excel(self.file_path, sheet_name=self.sheet_name, index=False)
    
    def parse_time_string(self, time_str):
        """Parse HH:MM string to Excel time format"""
        if pd.isna(time_str) or time_str is None or time_str == '':
            return ''
        
        try:
            # If it's already a datetime.time object
            if isinstance(time_str, dt_time):
                return time_str
            
            # If it's a string in HH:MM format
            if isinstance(time_str, str):
                # Handle Excel serial time (like 0.5 for 12:00)
                if '.' in time_str:
                    try:
                        # Convert Excel serial time to datetime.time
                        excel_time = float(time_str)
                        hours = int(excel_time * 24)
                        minutes = int((excel_time * 24 * 60) % 60)
                        return dt_time(hours, minutes)
                    except:
                        return ''
                
                # Parse HH:MM string
                if ':' in time_str:
                    parts = time_str.split(':')
                    if len(parts) >= 2:
                        hours = int(parts[0])
                        minutes = int(parts[1])
                        return dt_time(hours, minutes)
            
            return ''
        except Exception as e:
            print(f"⚠️  Error parsing time '{time_str}': {e}")
            return ''
    
    def format_time(self, datetime_obj):
        """Format datetime to HH:MM format"""
        if pd.isna(datetime_obj) or datetime_obj is None:
            return ''
        
        try:
            if isinstance(datetime_obj, str):
                # Try to parse string
                try:
                    datetime_obj = pd.to_datetime(datetime_obj)
                except:
                    # Check if it's already in HH:MM format
                    if ':' in datetime_obj:
                        return self.parse_time_string(datetime_obj)
                    return ''
            
            # Convert to datetime.time object
            if hasattr(datetime_obj, 'time'):
                return datetime_obj.time()
            elif hasattr(datetime_obj, 'hour'):
                return dt_time(datetime_obj.hour, datetime_obj.minute)
            else:
                # Try to extract time from string
                time_str = str(datetime_obj)
                if ':' in time_str:
                    parts = time_str.split(':')
                    if len(parts) >= 2:
                        hours = int(float(parts[0]))
                        minutes = int(float(parts[1]))
                        return dt_time(hours, minutes)
        except Exception as e:
            print(f"⚠️  Error formatting time '{datetime_obj}': {e}")
        
        return ''
    
    def format_date(self, datetime_obj):
        """Format datetime to YYYY-MM-DD format"""
        if pd.isna(datetime_obj) or datetime_obj is None:
            return ''
        
        try:
            if isinstance(datetime_obj, str):
                try:
                    datetime_obj = pd.to_datetime(datetime_obj)
                except:
                    return ''
            
            if hasattr(datetime_obj, 'strftime'):
                return datetime_obj.strftime('%Y-%m-%d')
            elif hasattr(datetime_obj, 'year'):
                return f"{datetime_obj.year}-{datetime_obj.month:02d}-{datetime_obj.day:02d}"
        except Exception as e:
            print(f"⚠️  Error formatting date '{datetime_obj}': {e}")
        
        return ''
    
    def add_job(self, job):
        """Add a new job to the spreadsheet - REAL-TIME UPDATE"""
        try:
            # Ensure file exists
            if not os.path.exists(self.file_path):
                self.initialize_spreadsheet()
            
            # Load existing data
            try:
                df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            except:
                # If file exists but is empty/corrupt, reinitialize
                self.initialize_spreadsheet()
                df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            # Format times for spreadsheet
            creation_date = self.format_date(job.created_time)
            start_time = self.format_time(job.start_time)
            end_time = self.format_time(job.completion_time)
            
            # Create new row WITH CORRECT ORDER
            new_row = {
                'Job Number': int(job.job_number),
                'Creation Date': creation_date,
                'Start Time': start_time,
                'End Time': end_time,
                'Room Number': job.room_number,
                'Issue': job.issue[:1000] if job.issue else '',  # Limit but allow longer text
                'Resolution': job.resolution[:1000] if job.resolution else ''  # Limit but allow longer text
            }
            
            # Append new row
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
            # Save back to Excel with formatting - REAL-TIME SAVE
            self.save_with_formatting(df)
            
            print(f"📝 Job {job.job_number} ADDED to spreadsheet in real-time")
            print(f"   Order: Job #{job.job_number}, Date: {creation_date}, Start: {start_time}, "
                  f"End: {end_time}, Room: {job.room_number}")
            return True
            
        except Exception as e:
            print(f"❌ Error adding job to spreadsheet: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_job(self, job):
        """Update an existing job in the spreadsheet - REAL-TIME UPDATE"""
        try:
            # Ensure file exists
            if not os.path.exists(self.file_path):
                print(f"⚠️  Spreadsheet doesn't exist, creating it...")
                self.initialize_spreadsheet()
                # Since this is an update but file didn't exist, add it as new
                return self.add_job(job)
            
            # Load existing data
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            # Find the job
            job_num_int = int(job.job_number)
            mask = df['Job Number'] == job_num_int
            
            if mask.any():
                # Format times for spreadsheet
                creation_date = self.format_date(job.created_time)
                start_time = self.format_time(job.start_time)
                end_time = self.format_time(job.completion_time)
                
                print(f"🔄 Updating job {job.job_number} in spreadsheet...")
                print(f"   Creation Date: {creation_date}")
                print(f"   Start Time: {start_time}")
                print(f"   End Time: {end_time}")
                print(f"   Status: {job.status.value}")
                print(f"   Resolution: {job.resolution[:50] if job.resolution else 'None'}")
                
                # Update ALL fields - REAL-TIME UPDATE
                df.loc[mask, 'Creation Date'] = creation_date
                df.loc[mask, 'Start Time'] = start_time
                df.loc[mask, 'End Time'] = end_time
                df.loc[mask, 'Room Number'] = job.room_number
                df.loc[mask, 'Issue'] = job.issue[:1000] if job.issue else ''
                df.loc[mask, 'Resolution'] = job.resolution[:1000] if job.resolution else ''
                
                # Save back to Excel with formatting - REAL-TIME SAVE
                self.save_with_formatting(df)
                
                print(f"✅ Job {job.job_number} UPDATED in spreadsheet in real-time")
                print(f"   Updated: Date: {creation_date}, Start: {start_time}, "
                      f"End: {end_time}, Room: {job.room_number}")
                return True
            else:
                # Job not found, add it as new
                print(f"⚠️  Job {job.job_number} not found in spreadsheet, adding as new...")
                return self.add_job(job)
            
        except Exception as e:
            print(f"❌ Error updating job in spreadsheet: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_job_resolution_only(self, job_number, resolution_text):
        """Update only the resolution for a job - REAL-TIME"""
        try:
            if not os.path.exists(self.file_path):
                print(f"⚠️  Spreadsheet doesn't exist for resolution update")
                return False
            
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            job_num_int = int(job_number)
            mask = df['Job Number'] == job_num_int
            
            if mask.any():
                df.loc[mask, 'Resolution'] = resolution_text[:1000]
                
                # Save back to Excel with formatting
                self.save_with_formatting(df)
                
                print(f"✅ Resolution UPDATED for job {job_number} in real-time")
                return True
            
            print(f"⚠️  Job {job_number} not found in spreadsheet for resolution update")
            return False
            
        except Exception as e:
            print(f"❌ Error updating resolution in spreadsheet: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_weekly_jobs(self, week_start_date=None):
        """Get all jobs for a specific week"""
        try:
            if not os.path.exists(self.file_path):
                return pd.DataFrame()
            
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            if df.empty:
                return pd.DataFrame()
            
            # Convert date columns to datetime
            if 'Creation Date' in df.columns:
                df['Creation Date'] = pd.to_datetime(df['Creation Date'], errors='coerce')
            
            if week_start_date is None:
                # Get last week's jobs (Monday to Sunday)
                today = datetime.now()
                # Find most recent Monday
                days_since_monday = today.weekday()
                last_monday = today - pd.Timedelta(days=days_since_monday + 7)
                last_sunday = last_monday + pd.Timedelta(days=6)
            else:
                last_monday = pd.Timestamp(week_start_date)
                last_sunday = last_monday + pd.Timedelta(days=6)
            
            # Filter jobs from last week
            weekly_jobs = df[
                (df['Creation Date'] >= last_monday) & 
                (df['Creation Date'] <= last_sunday)
            ]
            
            return weekly_jobs
            
        except Exception as e:
            print(f"❌ Error getting weekly jobs: {e}")
            return pd.DataFrame()
    
    def get_all_jobs(self):
        """Get all jobs from spreadsheet"""
        try:
            if not os.path.exists(self.file_path):
                return pd.DataFrame()
            
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            return df
            
        except Exception as e:
            print(f"❌ Error getting all jobs: {e}")
            return pd.DataFrame()
    
    def get_job_count(self):
        """Get the total number of jobs in spreadsheet"""
        try:
            if not os.path.exists(self.file_path):
                return 0
            
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            return len(df)
            
        except Exception as e:
            print(f"❌ Error getting job count: {e}")
            return 0
    
    def debug_spreadsheet(self):
        """Debug function to print spreadsheet contents"""
        try:
            if not os.path.exists(self.file_path):
                print("❌ Spreadsheet file doesn't exist")
                return
            
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            print(f"\n📊 SPREADSHEET DEBUG INFO")
            print(f"=" * 50)
            print(f"File: {self.file_path}")
            print(f"Rows: {len(df)}")
            print(f"Columns: {list(df.columns)}")
            print(f"\nColumn dtypes:")
            for col in df.columns:
                print(f"  {col}: {df[col].dtype}")
            print(f"\nFirst 5 rows:")
            print(df.head())
            print(f"\nLast 5 rows:")
            print(df.tail())
            print(f"=" * 50)
            
            # Check for specific job numbers
            if 'Job Number' in df.columns:
                print(f"\n🔍 Checking job completion status:")
                for job_num in df['Job Number'].unique()[:10]:  # First 10 unique jobs
                    job_data = df[df['Job Number'] == job_num]
                    if not job_data.empty:
                        end_time = job_data.iloc[0]['End Time']
                        resolution = job_data.iloc[0]['Resolution']
                        print(f"   Job #{job_num}: End Time='{end_time}' (type: {type(end_time)}), Resolution='{resolution[:30]}...'")
            
        except Exception as e:
            print(f"❌ Error debugging spreadsheet: {e}")
            
    def remove_job(self, job_number):
        """Remove a job from the spreadsheet"""
        try:
            if not os.path.exists(self.file_path):
                print(f"⚠️  Spreadsheet doesn't exist for job removal")
                return False
            
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            job_num_int = int(job_number)
            initial_count = len(df)
            
            # Remove the job
            df = df[df['Job Number'] != job_num_int]
            
            if len(df) < initial_count:
                # Save back to Excel with formatting
                self.save_with_formatting(df)
                print(f"✅ Job #{job_number} removed from spreadsheet")
                return True
            else:
                print(f"⚠️  Job #{job_number} not found in spreadsheet")
                return False
                
        except Exception as e:
            print(f"❌ Error removing job from spreadsheet: {e}")
            import traceback
            traceback.print_exc()
            return False
